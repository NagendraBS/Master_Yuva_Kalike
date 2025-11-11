import os
from openpyxl import load_workbook

# Function to determine Parameter2 and Parameter3 based on Parameter1
def get_parameters(AuditScope):
    if AuditScope == 'SA3':
        return 'Application Layer', 'Semi Annual'
    elif AuditScope == 'SA5':
        return 'Platform or OS Layer', 'Quarterly'
    elif AuditScope == 'SA11':
        return 'Database Layer', 'Quarterly'
    else:
        return None, None  # Default case if no match

# Path to the Excel file
excel_file_path = r'C:\Users\nmurthy\OneDrive - GalaxE. Solutions, Inc\Documents\POC Entitlement validation\Read_Scopes_Sample.xlsx'

# Check if the file exists
if os.path.exists(excel_file_path):
    # Load the workbook and select the active sheet
    wb = load_workbook(excel_file_path)
    sheet = wb.active

    # Iterate over rows in the sheet (assuming data starts from row 2)
    for row in range(2, sheet.max_row + 1):
        AuditScope = sheet.cell(row=row, column=1).value  # Assuming Parameter1 is in Column A

        # Get corresponding Parameter2 and Parameter3
        Review Layer, IAR Frequency = get_parameters(AuditScope)

        # Write the values back to the sheet if they are not None
        if Review Layer is not None and IAR Frequency is not None:
            sheet.cell(row=row, column=2).value = Review Layer  # Write to Column B
            sheet.cell(row=row, column=3).value = IAR Frequency  # Write to Column C

    # Save the changes
    wb.save(excel_file_path)
    print("Parameters updated successfully.")
else:
    print("The specified Excel file does not exist.")
