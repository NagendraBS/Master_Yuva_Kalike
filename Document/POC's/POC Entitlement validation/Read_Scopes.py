import os
from openpyxl import load_workbook

# Function to determine Parameter2 and Parameter3 based on Parameter1
def get_parameters(AuditScope):
    if AuditScope == 'SA3':
        return 'Application Layer', 'Semi Annual'
    elif AuditScope == 'SA5':
        return 'Platform or OS Layer', 'Quarterly'
    elif AuditScope == 'SA11':
        return 'Database Layer', 'Quarterly'
    else:
        return None, None  # Default case if no match

l# Path to the Excel file
excel_file_path = r'C:\Users\nmurthy\Documents\AuditScope\Read_Scopes_Sample.xlsx'

# Check if the file exists
if os.path.exists(excel_file_path):
    # Load the workbook and select the active sheet
    wb = load_workbook(excel_file_path)
    sheet = wb.active

    # Iterate over rows in the sheet (assuming data starts from row 2)
    for row in range(2, sheet.max_row + 1):
        AuditScope = sheet.cell(row=row, column=1).value  # Assuming Parameter1 is in Column A

        # Get corresponding Parameter2 and Parameter3
        ReviewLayer, IARFrequency = get_parameters(AuditScope)

        # Write the values back to the sheet if they are not None
        if ReviewLayer is not None and IARFrequency is not None:
            sheet.cell(row=row, column=2).value = ReviewLayer  # Write to Column B
            sheet.cell(row=row, column=3).value = IARFrequency  # Write to Column C

    # Save the changes
    wb.save(excel_file_path)
    print("Parameters updated successfully.")
else:
    print("The specified Excel file does not exist.")
